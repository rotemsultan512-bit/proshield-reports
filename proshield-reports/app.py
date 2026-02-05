from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_from_directory
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from PIL import Image
from datetime import datetime
from io import BytesIO
from sqlalchemy import or_
import os
import uuid
import json

from config import Config
from models import db, User, Report, ReportProduct, ReportImage, ReportDocument, CompanyProject, InventoryItem, InventoryTransaction, PRODUCTS

app = Flask(__name__)
app.config.from_object(Config)

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'יש להתחבר כדי לגשת לעמוד זה'


def _is_production_runtime() -> bool:
    """Detect production runtime (Render or any environment with DATABASE_URL).

    We use this to auto-initialize the DB on startup so the first login doesn't
    crash with missing tables.
    """
    return (
        os.environ.get('RENDER', '').lower() == 'true'
        or bool(os.environ.get('RENDER_SERVICE_ID'))
        or bool(os.environ.get('DATABASE_URL'))
    )


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def init_db():
    """Initialize database and create default admin user.

    In production (Render), the SQLite DB may be created in /tmp and start empty.
    Without this, the first login attempt can crash because tables don't exist.
    """
    with app.app_context():
        db.create_all()

        # Lightweight schema migrations (SQLite only)
        # NOTE: db.create_all() does not add new columns to existing tables.
        from sqlalchemy import text
        try:
            if db.engine.dialect.name == 'sqlite':
                report_columns = [
                    row[1] for row in db.session.execute(text('PRAGMA table_info(reports)')).fetchall()
                ]

                if 'customer_name' not in report_columns:
                    db.session.execute(text('ALTER TABLE reports ADD COLUMN customer_name VARCHAR(200)'))

                if 'company_project' not in report_columns:
                    db.session.execute(text('ALTER TABLE reports ADD COLUMN company_project VARCHAR(200)'))

                if 'installation_type' not in report_columns:
                    db.session.execute(text('ALTER TABLE reports ADD COLUMN installation_type VARCHAR(500)'))

                if 'installation_types' not in report_columns:
                    db.session.execute(text('ALTER TABLE reports ADD COLUMN installation_types TEXT'))

                if 'protections_count' not in report_columns:
                    db.session.execute(text('ALTER TABLE reports ADD COLUMN protections_count INTEGER'))

                # Company/Project list table
                db.session.execute(text('''
                    CREATE TABLE IF NOT EXISTS company_projects (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name VARCHAR(200) UNIQUE NOT NULL,
                        is_active BOOLEAN DEFAULT 1,
                        created_at DATETIME
                    )
                '''))

                # Inventory tables
                db.session.execute(text('''
                    CREATE TABLE IF NOT EXISTS inventory_items (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        product_name VARCHAR(200) UNIQUE NOT NULL,
                        quantity_unit FLOAT DEFAULT 0,
                        quantity_meter FLOAT DEFAULT 0,
                        updated_at DATETIME
                    )
                '''))
                db.session.execute(text('''
                    CREATE TABLE IF NOT EXISTS inventory_transactions (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        product_name VARCHAR(200) NOT NULL,
                        change_type VARCHAR(50) NOT NULL,
                        quantity FLOAT NOT NULL,
                        unit VARCHAR(20) NOT NULL,
                        report_id INTEGER,
                        user_id INTEGER,
                        notes VARCHAR(500),
                        created_at DATETIME,
                        FOREIGN KEY(report_id) REFERENCES reports(id),
                        FOREIGN KEY(user_id) REFERENCES users(id)
                    )
                '''))

                # Report products: quantity unit
                product_columns = [
                    row[1] for row in db.session.execute(text('PRAGMA table_info(report_products)')).fetchall()
                ]
                if 'quantity_unit' not in product_columns:
                    db.session.execute(text('ALTER TABLE report_products ADD COLUMN quantity_unit VARCHAR(20)'))

                # Rename PP Tape -> לוח PP מ"מ 4 (existing data)
                db.session.execute(text("UPDATE inventory_items SET product_name = 'לוח PP מ\"מ 4' WHERE product_name = 'PP Tape'"))
                db.session.execute(text("UPDATE report_products SET product_name = 'לוח PP מ\"מ 4' WHERE product_name = 'PP Tape'"))
                db.session.execute(text("UPDATE inventory_transactions SET product_name = 'לוח PP מ\"מ 4' WHERE product_name = 'PP Tape'"))

                db.session.commit()

            elif db.engine.dialect.name == 'postgresql':
                # Postgres supports IF NOT EXISTS
                db.session.execute(text('ALTER TABLE reports ADD COLUMN IF NOT EXISTS customer_name VARCHAR(200)'))
                db.session.execute(text('ALTER TABLE reports ADD COLUMN IF NOT EXISTS company_project VARCHAR(200)'))
                db.session.execute(text('ALTER TABLE reports ADD COLUMN IF NOT EXISTS installation_type VARCHAR(500)'))
                db.session.execute(text('ALTER TABLE reports ADD COLUMN IF NOT EXISTS installation_types TEXT'))
                db.session.execute(text('ALTER TABLE reports ADD COLUMN IF NOT EXISTS protections_count INTEGER'))
                db.session.execute(text('ALTER TABLE report_products ADD COLUMN IF NOT EXISTS quantity_unit VARCHAR(20)'))

                db.session.execute(text('''
                    CREATE TABLE IF NOT EXISTS company_projects (
                        id SERIAL PRIMARY KEY,
                        name VARCHAR(200) UNIQUE NOT NULL,
                        is_active BOOLEAN DEFAULT TRUE,
                        created_at TIMESTAMP
                    )
                '''))

                db.session.execute(text('''
                    CREATE TABLE IF NOT EXISTS inventory_items (
                        id SERIAL PRIMARY KEY,
                        product_name VARCHAR(200) UNIQUE NOT NULL,
                        quantity_unit DOUBLE PRECISION DEFAULT 0,
                        quantity_meter DOUBLE PRECISION DEFAULT 0,
                        updated_at TIMESTAMP
                    )
                '''))
                db.session.execute(text('''
                    CREATE TABLE IF NOT EXISTS inventory_transactions (
                        id SERIAL PRIMARY KEY,
                        product_name VARCHAR(200) NOT NULL,
                        change_type VARCHAR(50) NOT NULL,
                        quantity DOUBLE PRECISION NOT NULL,
                        unit VARCHAR(20) NOT NULL,
                        report_id INTEGER,
                        user_id INTEGER,
                        notes VARCHAR(500),
                        created_at TIMESTAMP
                    )
                '''))

                db.session.execute(text("UPDATE inventory_items SET product_name = 'לוח PP מ\"מ 4' WHERE product_name = 'PP Tape'"))
                db.session.execute(text("UPDATE report_products SET product_name = 'לוח PP מ\"מ 4' WHERE product_name = 'PP Tape'"))
                db.session.execute(text("UPDATE inventory_transactions SET product_name = 'לוח PP מ\"מ 4' WHERE product_name = 'PP Tape'"))

                db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"Schema migration skipped/failed: {e}")

        # Ensure inventory items exist for all products
        existing_items = {i.product_name for i in InventoryItem.query.all()}
        for product in PRODUCTS:
            if product not in existing_items:
                db.session.add(InventoryItem(product_name=product, quantity_unit=0, quantity_meter=0))
        db.session.commit()

        # Create default admin if not exists
        admin = User.query.filter_by(username='rotem').first()
        if not admin:
            admin = User(
                username='rotem',
                role='admin',
                full_name='רותם'
            )
            admin.set_password('proshield2025')
            db.session.add(admin)
            db.session.commit()
            print("Default admin user created: rotem / proshield2025")


# Auto-init DB on production startup (Gunicorn imports app.py but does not run __main__).
# Safe to call multiple times because create_all() is idempotent.
if _is_production_runtime():
    init_db()

def allowed_file(filename, file_type='image'):
    if file_type == 'image':
        allowed = Config.ALLOWED_IMAGE_EXTENSIONS
    else:
        allowed = Config.ALLOWED_DOCUMENT_EXTENSIONS
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed

def compress_image(image_data, max_size=Config.MAX_IMAGE_DIMENSION, quality=Config.JPEG_QUALITY):
    """Compress and resize image"""
    img = Image.open(BytesIO(image_data))

    # Convert to RGB if necessary
    if img.mode in ('RGBA', 'P'):
        img = img.convert('RGB')

    # Resize if too large
    if img.width > max_size or img.height > max_size:
        img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)

    # Save to bytes
    output = BytesIO()
    img.save(output, format='JPEG', quality=quality, optimize=True)
    output.seek(0)
    return output

def save_file(file, report_id, file_type='image'):
    """Save uploaded file with compression for images"""
    if not file:
        return None

    # Create directory structure
    report_dir = os.path.join(Config.UPLOAD_FOLDER, str(report_id))
    if file_type == 'image':
        save_dir = os.path.join(report_dir, 'images')
    elif file_type == 'delivery_note':
        save_dir = os.path.join(report_dir, 'delivery_note')
    else:
        save_dir = os.path.join(report_dir, 'documents')

    os.makedirs(save_dir, exist_ok=True)

    # Generate unique filename
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
    if file_type == 'image':
        ext = 'jpg'  # Always save images as JPEG after compression
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(save_dir, filename)

    if file_type == 'image':
        # Compress and save image
        compressed = compress_image(file.read())
        with open(filepath, 'wb') as f:
            f.write(compressed.read())
    else:
        file.save(filepath)

    # Return relative path for storage
    if file_type == 'image':
        subdir = 'images'
    elif file_type == 'delivery_note':
        subdir = 'delivery_note'
    else:
        subdir = 'documents'
    return os.path.join(str(report_id), subdir, filename)


def apply_inventory_change(product_name, quantity, unit, change_type, report_id=None, user_id=None, notes=None):
    """Apply inventory change and record transaction. Allows negative stock."""
    if unit not in ['unit', 'meter']:
        unit = 'unit'

    item = InventoryItem.query.filter_by(product_name=product_name).first()
    if not item:
        item = InventoryItem(product_name=product_name, quantity_unit=0, quantity_meter=0)
        db.session.add(item)
        db.session.flush()

    if unit == 'meter':
        item.quantity_meter = (item.quantity_meter or 0) + float(quantity)
    else:
        item.quantity_unit = (item.quantity_unit or 0) + float(quantity)

    tx = InventoryTransaction(
        product_name=product_name,
        change_type=change_type,
        quantity=float(quantity),
        unit=unit,
        report_id=report_id,
        user_id=user_id,
        notes=notes
    )
    db.session.add(tx)

# ============ ROUTES ============

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username', '').strip()
        password = data.get('password', '')

        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password) and user.is_active:
            login_user(user, remember=True)
            if request.is_json:
                return jsonify({'success': True, 'redirect': url_for('dashboard')})
            return redirect(url_for('dashboard'))

        error_msg = 'שם משתמש או סיסמה שגויים'
        if request.is_json:
            return jsonify({'success': False, 'error': error_msg}), 401
        flash(error_msg, 'error')

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/report/new')
@login_required
def new_report():
    return render_template('new_report.html', products=PRODUCTS)

@app.route('/report/<int:report_id>')
@login_required
def view_report(report_id):
    report = Report.query.get_or_404(report_id)

    # Check access permission
    if not current_user.is_admin() and report.user_id != current_user.id:
        flash('אין לך הרשאה לצפות בדוח זה', 'error')
        return redirect(url_for('dashboard'))

    return render_template('view_report.html', report=report)

@app.route('/report/<int:report_id>/edit')
@login_required
def edit_report(report_id):
    report = Report.query.get_or_404(report_id)

    if not current_user.is_admin() and report.user_id != current_user.id:
        flash('אין לך הרשאה לערוך דוח זה', 'error')
        return redirect(url_for('dashboard'))

    return render_template('new_report.html', products=PRODUCTS, edit_mode=True, report_id=report.id)

@app.route('/settings')
@login_required
def settings():
    return render_template('settings.html')

@app.route('/admin')
@login_required
def admin_dashboard():
    if not current_user.is_admin():
        flash('אין לך הרשאה לגשת לעמוד זה', 'error')
        return redirect(url_for('dashboard'))
    return render_template('admin.html')

@app.route('/inventory')
@login_required
def inventory_page():
    if not current_user.is_admin():
        flash('אין לך הרשאה לגשת לעמוד זה', 'error')
        return redirect(url_for('dashboard'))
    return render_template('inventory.html')

# ============ API ROUTES ============

@app.route('/api/reports', methods=['GET'])
@login_required
def get_reports():
    """Get reports - all for admin, own for regular users"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    # Filters
    report_type = request.args.get('type')
    status = request.args.get('status')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    search = request.args.get('search')
    user_search = request.args.get('user_search')
    user_id = request.args.get('user_id', type=int)

    # Base query
    if current_user.is_admin():
        query = Report.query
        if user_id:
            query = query.filter(Report.user_id == user_id)
    else:
        query = Report.query.filter_by(user_id=current_user.id)

    # Apply filters
    if report_type:
        query = query.filter(Report.report_type == report_type)
    if status:
        query = query.filter(Report.status == status)
    if date_from:
        query = query.filter(Report.timestamp >= datetime.fromisoformat(date_from))
    if date_to:
        query = query.filter(Report.timestamp <= datetime.fromisoformat(date_to + 'T23:59:59'))
    if search:
        query = query.filter(or_(
            Report.address.ilike(f'%{search}%'),
            Report.customer_name.ilike(f'%{search}%'),
            Report.company_project.ilike(f'%{search}%')
        ))

    if user_search and current_user.is_admin():
        query = query.join(User).filter(or_(
            User.full_name.ilike(f'%{user_search}%'),
            User.username.ilike(f'%{user_search}%')
        ))

    # Order and paginate
    query = query.order_by(Report.timestamp.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        'reports': [r.to_dict() for r in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': page
    })

@app.route('/api/reports/stats', methods=['GET'])
@login_required
def get_reports_stats():
    """Get report statistics for dashboard"""
    from sqlalchemy import func, extract

    # Base query based on user role
    if current_user.is_admin():
        base_query = Report.query
    else:
        base_query = Report.query.filter_by(user_id=current_user.id)

    # Total reports
    total = base_query.count()

    # Reports by type
    delivery = base_query.filter(Report.report_type == 'delivery').count()
    installation = base_query.filter(Report.report_type == 'installation').count()

    # Reports this month
    now = datetime.utcnow()
    first_day_of_month = datetime(now.year, now.month, 1)
    this_month = base_query.filter(Report.timestamp >= first_day_of_month).count()

    return jsonify({
        'total': total,
        'delivery': delivery,
        'installation': installation,
        'this_month': this_month
    })

@app.route('/api/reports', methods=['POST'])
@login_required
def create_report():
    """Create a new report"""
    try:
        # Get form data
        report_type = request.form.get('report_type')
        address = request.form.get('address')
        status = request.form.get('status')
        notes = request.form.get('notes', '')
        products_json = request.form.get('products', '[]')

        # Delivery / Installation extra fields
        customer_name = (request.form.get('customer_name') or '').strip()
        company_project = (request.form.get('company_project') or '').strip()
        report_datetime_raw = request.form.get('report_datetime')

        # installation_types is expected to be a JSON array string from the client
        installation_types_raw = request.form.get('installation_types')
        # Backward compatibility (if a client still sends a single installation_type)
        installation_type_single = request.form.get('installation_type')

        protections_count_raw = request.form.get('protections_count')

        # Validate required fields
        if not all([report_type, address, status]):
            return jsonify({'success': False, 'error': 'יש למלא את כל השדות הנדרשים'}), 400

        if not customer_name:
            return jsonify({'success': False, 'error': 'יש להזין שם לקוח'}), 400

        # Parse report date/time
        report_timestamp = None
        if report_datetime_raw:
            try:
                # Expecting "YYYY-MM-DDTHH:MM" from datetime-local
                report_timestamp = datetime.fromisoformat(report_datetime_raw)
            except ValueError:
                return jsonify({'success': False, 'error': 'תאריך/שעה לא תקין'}), 400
        else:
            return jsonify({'success': False, 'error': 'יש לבחור תאריך ושעה'}), 400

        # Installation extra validation
        protections_count = None
        installation_types_list = []
        installation_type_display = None

        if report_type == 'installation':
            # Parse types
            if installation_types_raw:
                try:
                    installation_types_list = json.loads(installation_types_raw)
                    if not isinstance(installation_types_list, list):
                        installation_types_list = []
                except json.JSONDecodeError:
                    installation_types_list = []

            # Fallback to legacy single field
            if not installation_types_list and installation_type_single:
                installation_types_list = [installation_type_single]

            # Validate
            if not installation_types_list:
                return jsonify({'success': False, 'error': 'יש לבחור לפחות סוג התקנה אחד'}), 400

            installation_type_display = ', '.join([str(x) for x in installation_types_list if x])

            try:
                protections_count = int(protections_count_raw) if protections_count_raw is not None else None
            except (TypeError, ValueError):
                protections_count = None

        # Parse products
        try:
            products_data = json.loads(products_json)
        except json.JSONDecodeError:
            products_data = []

        if not products_data:
            return jsonify({'success': False, 'error': 'יש לבחור לפחות מוצר אחד'}), 400

        # Create report
        report = Report(
            user_id=current_user.id,
            report_type=report_type,
            customer_name=customer_name,
            company_project=company_project or None,
            installation_type=installation_type_display if report_type == 'installation' else None,
            installation_types=(
                json.dumps(installation_types_list, ensure_ascii=False)
                if report_type == 'installation'
                else None
            ),
            protections_count=protections_count if report_type == 'installation' else None,
            address=address,
            status=status,
            notes=notes,
            timestamp=report_timestamp
        )
        db.session.add(report)
        db.session.flush()  # Get report ID

        # Add products
        for product in products_data:
            if product.get('name') and product.get('quantity'):
                unit = product.get('unit') or 'unit'
                if unit not in ['unit', 'meter']:
                    unit = 'unit'

                report_product = ReportProduct(
                    report_id=report.id,
                    product_name=product['name'],
                    quantity=float(product['quantity']),
                    quantity_unit=unit
                )
                db.session.add(report_product)

                # Inventory: subtract reported quantity
                apply_inventory_change(
                    product_name=product['name'],
                    quantity=-float(product['quantity']),
                    unit=unit,
                    change_type='report',
                    report_id=report.id,
                    user_id=current_user.id,
                    notes=f"Report #{report.id} (offline sync)"
                )

                # Inventory: subtract reported quantity
                apply_inventory_change(
                    product_name=product['name'],
                    quantity=-float(product['quantity']),
                    unit=unit,
                    change_type='report',
                    report_id=report.id,
                    user_id=current_user.id,
                    notes=f"Report #{report.id}"
                )

        # Handle delivery note upload (MANDATORY for delivery reports)
        if report_type == 'delivery':
            delivery_note = request.files.get('delivery_note')
            if not delivery_note or not delivery_note.filename:
                db.session.rollback()
                return jsonify({'success': False, 'error': 'יש להעלות תעודת משלוח חתומה עבור דוח אספקה'}), 400

            if not allowed_file(delivery_note.filename, 'document'):
                db.session.rollback()
                return jsonify({'success': False, 'error': 'סוג קובץ לא חוקי. יש להעלות PDF או תמונה'}), 400

            # Check file size (10MB max)
            delivery_note.seek(0, 2)  # Seek to end
            file_size = delivery_note.tell()
            delivery_note.seek(0)  # Reset to beginning

            if file_size > Config.MAX_DOCUMENT_SIZE:
                db.session.rollback()
                return jsonify({'success': False, 'error': 'קובץ תעודת המשלוח גדול מדי (מקסימום 10MB)'}), 400

            doc_path = save_file(delivery_note, report.id, 'delivery_note')
            if doc_path:
                report_doc = ReportDocument(
                    report_id=report.id,
                    document_path=doc_path,
                    original_filename=secure_filename(delivery_note.filename)
                )
                db.session.add(report_doc)

        # Handle image uploads
        images = request.files.getlist('images')
        for image in images:
            if image and image.filename and allowed_file(image.filename, 'image'):
                if image.content_length and image.content_length > Config.MAX_IMAGE_SIZE:
                    continue  # Skip oversized files

                image_path = save_file(image, report.id, 'image')
                if image_path:
                    image_type = 'goods' if report_type == 'delivery' else 'project'
                    report_image = ReportImage(
                        report_id=report.id,
                        image_path=image_path,
                        image_type=image_type
                    )
                    db.session.add(report_image)

        db.session.commit()

        return jsonify({
            'success': True,
            'report_id': report.id,
            'message': 'הדוח נשמר בהצלחה'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'שגיאה בשמירת הדוח: {str(e)}'}), 500


@app.route('/api/reports/<int:report_id>', methods=['GET'])
@login_required
def get_report(report_id):
    """Get single report"""
    report = Report.query.get_or_404(report_id)

    if not current_user.is_admin() and report.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    return jsonify(report.to_dict())


@app.route('/api/reports/<int:report_id>', methods=['PUT'])
@login_required
def update_report(report_id):
    """Update a report and sync inventory"""
    report = Report.query.get_or_404(report_id)

    if not current_user.is_admin() and report.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    try:
        report_type = request.form.get('report_type')
        address = request.form.get('address')
        status = request.form.get('status')
        notes = request.form.get('notes', '')
        products_json = request.form.get('products', '[]')

        customer_name = (request.form.get('customer_name') or '').strip()
        company_project = (request.form.get('company_project') or '').strip()
        report_datetime_raw = request.form.get('report_datetime')

        installation_types_raw = request.form.get('installation_types')
        installation_type_single = request.form.get('installation_type')

        if not all([report_type, address, status]):
            return jsonify({'success': False, 'error': 'יש למלא את כל השדות הנדרשים'}), 400

        if not customer_name:
            return jsonify({'success': False, 'error': 'יש להזין שם לקוח'}), 400

        # Parse report date/time
        report_timestamp = None
        if report_datetime_raw:
            try:
                report_timestamp = datetime.fromisoformat(report_datetime_raw)
            except ValueError:
                return jsonify({'success': False, 'error': 'תאריך/שעה לא תקין'}), 400
        else:
            return jsonify({'success': False, 'error': 'יש לבחור תאריך ושעה'}), 400

        # Installation types
        installation_types_list = []
        installation_type_display = None
        if report_type == 'installation':
            if installation_types_raw:
                try:
                    installation_types_list = json.loads(installation_types_raw)
                    if not isinstance(installation_types_list, list):
                        installation_types_list = []
                except json.JSONDecodeError:
                    installation_types_list = []

            if not installation_types_list and installation_type_single:
                installation_types_list = [installation_type_single]

            if not installation_types_list:
                return jsonify({'success': False, 'error': 'יש לבחור לפחות סוג התקנה אחד'}), 400

            installation_type_display = ', '.join([str(x) for x in installation_types_list if x])

        # Parse products
        try:
            products_data = json.loads(products_json)
        except json.JSONDecodeError:
            products_data = []

        if not products_data:
            return jsonify({'success': False, 'error': 'יש לבחור לפחות מוצר אחד'}), 400

        # Revert inventory from old products
        for old_product in report.products:
            apply_inventory_change(
                product_name=old_product.product_name,
                quantity=float(old_product.quantity),
                unit=old_product.quantity_unit or 'unit',
                change_type='report_edit',
                report_id=report.id,
                user_id=current_user.id,
                notes=f"Revert Report #{report.id}"
            )

        # Clear old products
        ReportProduct.query.filter_by(report_id=report.id).delete()

        # Update report fields
        report.report_type = report_type
        report.customer_name = customer_name
        report.company_project = company_project or None
        report.address = address
        report.status = status
        report.notes = notes
        report.timestamp = report_timestamp
        report.installation_types = (
            json.dumps(installation_types_list, ensure_ascii=False)
            if report_type == 'installation'
            else None
        )
        report.installation_type = installation_type_display if report_type == 'installation' else None

        # Add new products + apply inventory
        for product in products_data:
            if product.get('name') and product.get('quantity'):
                unit = product.get('unit') or 'unit'
                if unit not in ['unit', 'meter']:
                    unit = 'unit'

                report_product = ReportProduct(
                    report_id=report.id,
                    product_name=product['name'],
                    quantity=float(product['quantity']),
                    quantity_unit=unit
                )
                db.session.add(report_product)

                apply_inventory_change(
                    product_name=product['name'],
                    quantity=-float(product['quantity']),
                    unit=unit,
                    change_type='report_edit',
                    report_id=report.id,
                    user_id=current_user.id,
                    notes=f"Update Report #{report.id}"
                )

        db.session.commit()
        return jsonify({'success': True, 'report_id': report.id})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'שגיאה בעדכון הדוח: {str(e)}'}), 500


@app.route('/api/reports/<int:report_id>', methods=['DELETE'])
@login_required
def delete_report(report_id):
    """Delete a report"""
    report = Report.query.get_or_404(report_id)

    # Check permission
    if not current_user.is_admin() and report.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'אין לך הרשאה למחוק דוח זה'}), 403

    try:
        # Delete associated files
        report_dir = os.path.join(Config.UPLOAD_FOLDER, str(report_id))
        if os.path.exists(report_dir):
            import shutil
            shutil.rmtree(report_dir)

        db.session.delete(report)
        db.session.commit()

        return jsonify({'success': True, 'message': 'הדוח נמחק בהצלחה'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/user/password', methods=['POST'])
@login_required
def change_password():
    """Change user password"""
    data = request.get_json()
    current_password = data.get('current_password')
    new_password = data.get('new_password')

    if not current_user.check_password(current_password):
        return jsonify({'success': False, 'error': 'הסיסמה הנוכחית שגויה'}), 400

    if len(new_password) < 6:
        return jsonify({'success': False, 'error': 'הסיסמה החדשה חייבת להכיל לפחות 6 תווים'}), 400

    current_user.set_password(new_password)
    db.session.commit()

    return jsonify({'success': True, 'message': 'הסיסמה שונתה בהצלחה'})

@app.route('/api/users', methods=['GET'])
@login_required
def get_users():
    """Get all users (admin only)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    users = User.query.all()
    return jsonify({
        'users': [{
            'id': u.id,
            'username': u.username,
            'full_name': u.full_name,
            'role': u.role,
            'is_active': u.is_active
        } for u in users]
    })

@app.route('/api/users', methods=['POST'])
@login_required
def create_user():
    """Create new user (admin only)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    full_name = data.get('full_name', '').strip()
    role = data.get('role', 'user')

    if not all([username, password, full_name]):
        return jsonify({'success': False, 'error': 'יש למלא את כל השדות'}), 400

    if User.query.filter_by(username=username).first():
        return jsonify({'success': False, 'error': 'שם משתמש כבר קיים'}), 400

    user = User(username=username, full_name=full_name, role=role)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()

    return jsonify({'success': True, 'message': 'המשתמש נוצר בהצלחה'})

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
def update_user(user_id):
    """Update user details (admin only)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    data = request.get_json() or {}
    full_name = (data.get('full_name') or '').strip()
    role = (data.get('role') or 'user').strip()
    is_active = data.get('is_active')

    if role not in ['user', 'admin']:
        return jsonify({'success': False, 'error': 'תפקיד לא תקין'}), 400

    user = User.query.get_or_404(user_id)

    if user.id == current_user.id and role != 'admin':
        return jsonify({'success': False, 'error': 'לא ניתן להסיר הרשאת מנהל מעצמך'}), 400

    if full_name:
        user.full_name = full_name
    user.role = role

    if is_active is not None:
        user.is_active = bool(is_active)

    db.session.commit()
    return jsonify({'success': True, 'message': 'המשתמש עודכן בהצלחה'})

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
def delete_user(user_id):
    """Delete user (admin only)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    if user_id == current_user.id:
        return jsonify({'success': False, 'error': 'לא ניתן למחוק את עצמך'}), 400

    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()

    return jsonify({'success': True, 'message': 'המשתמש נמחק בהצלחה'})


@app.route('/api/company-projects', methods=['GET'])
@login_required
def get_company_projects():
    """Get company/project names"""
    include_inactive = request.args.get('include_inactive') == 'true'

    query = CompanyProject.query
    if not current_user.is_admin() or not include_inactive:
        query = query.filter(CompanyProject.is_active == True)  # noqa: E712

    projects = query.order_by(CompanyProject.name.asc()).all()
    return jsonify({
        'projects': [p.to_dict() for p in projects]
    })


@app.route('/api/company-projects', methods=['POST'])
@login_required
def create_company_project():
    """Create company/project (admin only)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'יש להזין שם'}), 400

    existing = CompanyProject.query.filter_by(name=name).first()
    if existing:
        existing.is_active = True
        db.session.commit()
        return jsonify({'success': True, 'project': existing.to_dict()})

    project = CompanyProject(name=name, is_active=True)
    db.session.add(project)
    db.session.commit()

    return jsonify({'success': True, 'project': project.to_dict()})


@app.route('/api/company-projects/<int:project_id>', methods=['DELETE'])
@login_required
def delete_company_project(project_id):
    """Delete company/project (admin only)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    project = CompanyProject.query.get_or_404(project_id)
    db.session.delete(project)
    db.session.commit()

    return jsonify({'success': True, 'message': 'הפריט נמחק בהצלחה'})


@app.route('/api/inventory', methods=['GET'])
@login_required
def get_inventory():
    """Get inventory list (admin only)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    # Ensure inventory items exist for all products
    existing_items = {i.product_name: i for i in InventoryItem.query.all()}
    for product in PRODUCTS:
        if product not in existing_items:
            item = InventoryItem(product_name=product, quantity_unit=0, quantity_meter=0)
            db.session.add(item)
            existing_items[product] = item
    db.session.commit()

    items = [existing_items[p].to_dict() for p in PRODUCTS if p in existing_items]
    return jsonify({'items': items})


@app.route('/api/inventory/adjust', methods=['POST'])
@login_required
def adjust_inventory():
    """Adjust inventory (admin only). Allows negative stock."""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    data = request.get_json() or {}
    items = data.get('items', [])
    if not isinstance(items, list):
        return jsonify({'success': False, 'error': 'נתונים לא תקינים'}), 400

    try:
        for item in items:
            product_name = (item.get('product_name') or '').strip()
            if not product_name:
                continue

            target_unit = float(item.get('quantity_unit') or 0)
            target_meter = float(item.get('quantity_meter') or 0)

            inv = InventoryItem.query.filter_by(product_name=product_name).first()
            if not inv:
                inv = InventoryItem(product_name=product_name, quantity_unit=0, quantity_meter=0)
                db.session.add(inv)
                db.session.flush()

            delta_unit = target_unit - (inv.quantity_unit or 0)
            delta_meter = target_meter - (inv.quantity_meter or 0)

            if delta_unit != 0:
                apply_inventory_change(
                    product_name=product_name,
                    quantity=delta_unit,
                    unit='unit',
                    change_type='adjustment',
                    user_id=current_user.id,
                    notes='Manual adjustment'
                )

            if delta_meter != 0:
                apply_inventory_change(
                    product_name=product_name,
                    quantity=delta_meter,
                    unit='meter',
                    change_type='adjustment',
                    user_id=current_user.id,
                    notes='Manual adjustment'
                )

        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


def _export_inventory_to_excel(items, transactions):
    from openpyxl import Workbook
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    ws_items = wb.active
    ws_items.title = "מלאי"

    ws_items.append(['מוצר', 'כמות יחידה', 'כמות מטר', 'עודכן לאחרונה'])
    for item in items:
        ws_items.append([
            item.product_name,
            item.quantity_unit or 0,
            item.quantity_meter or 0,
            item.updated_at.strftime('%d/%m/%Y %H:%M') if item.updated_at else ''
        ])

    ws_tx = wb.create_sheet(title="תנועות מלאי")
    ws_tx.append(['מוצר', 'סוג שינוי', 'כמות', 'יחידה', 'דוח', 'משתמש', 'הערה', 'תאריך'])
    for tx in transactions:
        ws_tx.append([
            tx.product_name,
            tx.change_type,
            tx.quantity,
            'יח׳' if tx.unit == 'unit' else 'מ׳',
            tx.report_id or '',
            tx.user_id or '',
            tx.notes or '',
            tx.created_at.strftime('%d/%m/%Y %H:%M') if tx.created_at else ''
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'inventory_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.route('/api/inventory/export')
@login_required
def export_inventory():
    """Export inventory and transactions (admin only)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    items = InventoryItem.query.order_by(InventoryItem.product_name.asc()).all()
    transactions = InventoryTransaction.query.order_by(InventoryTransaction.created_at.desc()).all()
    return _export_inventory_to_excel(items, transactions)


@app.route('/api/stats')
@login_required
def get_stats():
    """Get statistics (admin only)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    # Total reports
    total_reports = Report.query.count()
    completed_reports = Report.query.filter_by(status='completed').count()
    pending_reports = Report.query.filter_by(status='return_required').count()

    # Reports by type
    delivery_reports = Report.query.filter_by(report_type='delivery').count()
    installation_reports = Report.query.filter_by(report_type='installation').count()

    # Reports per user
    users = User.query.all()
    reports_per_user = []
    for user in users:
        count = Report.query.filter_by(user_id=user.id).count()
        reports_per_user.append({
            'user_id': user.id,
            'full_name': user.full_name,
            'count': count
        })

    return jsonify({
        'total_reports': total_reports,
        'completed_reports': completed_reports,
        'pending_reports': pending_reports,
        'delivery_reports': delivery_reports,
        'installation_reports': installation_reports,
        'reports_per_user': reports_per_user
    })

def _export_reports_to_excel(reports, filename_prefix):
    from openpyxl import Workbook
    from io import BytesIO
    from flask import send_file

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "דוחות"

    # Headers
    headers = [
        'מזהה',
        'משתמש',
        'סוג דוח',
        'שם לקוח',
        'חברת בניה/פרויקט',
        'סוגי התקנה',
        'כתובת',
        'סטטוס',
        'תאריך',
        'מוצרים',
        'הערות'
    ]
    ws.append(headers)

    # Data rows
    for report in reports:
        products_str = ', '.join([
            f"{p.product_name} ({p.quantity}{' מ׳' if p.quantity_unit == 'meter' else ' יח׳'})"
            for p in report.products
        ])
        ws.append([
            report.id,
            report.author.full_name if report.author else '',
            'אספקה' if report.report_type == 'delivery' else 'התקנה',
            report.customer_name or '',
            report.company_project or '',
            report.installation_type or '',
            report.address,
            'הושלם' if report.status == 'completed' else 'נדרש חזרה',
            report.timestamp.strftime('%d/%m/%Y %H:%M') if report.timestamp else '',
            products_str,
            report.notes or ''
        ])

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{filename_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.route('/api/export')
@login_required
def export_reports():
    """Export reports to Excel (admin only)"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'error': 'אין הרשאה'}), 403

    # Get filters
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    report_type = request.args.get('type')
    user_id = request.args.get('user_id', type=int)

    query = Report.query

    if user_id:
        query = query.filter(Report.user_id == user_id)
    if date_from:
        query = query.filter(Report.timestamp >= datetime.fromisoformat(date_from))
    if date_to:
        query = query.filter(Report.timestamp <= datetime.fromisoformat(date_to + 'T23:59:59'))
    if report_type:
        query = query.filter(Report.report_type == report_type)

    reports = query.order_by(Report.timestamp.desc()).all()

    return _export_reports_to_excel(reports, 'reports')


@app.route('/api/export/mine')
@login_required
def export_my_reports():
    """Export current user's reports (monthly by default)"""
    # Get filters
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    report_type = request.args.get('type')

    # Default: current month if no dates provided
    if not date_from and not date_to:
        now = datetime.utcnow()
        date_from = f"{now.year}-{now.month:02d}-01"
        date_to = f"{now.year}-{now.month:02d}-{now.day:02d}"

    query = Report.query.filter_by(user_id=current_user.id)

    if date_from:
        query = query.filter(Report.timestamp >= datetime.fromisoformat(date_from))
    if date_to:
        query = query.filter(Report.timestamp <= datetime.fromisoformat(date_to + 'T23:59:59'))
    if report_type:
        query = query.filter(Report.report_type == report_type)

    reports = query.order_by(Report.timestamp.desc()).all()

    return _export_reports_to_excel(reports, 'my_reports')

@app.route('/api/sync', methods=['POST'])
@login_required
def sync_offline_reports():
    """Sync offline reports"""
    data = request.get_json()
    offline_reports = data.get('reports', [])

    synced_count = 0
    errors = []

    for report_data in offline_reports:
        try:
            # Create report from offline data
            offline_type = report_data.get('report_type')
            offline_protections_raw = report_data.get('protections_count')
            try:
                offline_protections_count = int(offline_protections_raw) if offline_protections_raw is not None else None
            except (TypeError, ValueError):
                offline_protections_count = None

            offline_installation_types = report_data.get('installation_types')
            if isinstance(offline_installation_types, str):
                try:
                    offline_installation_types = json.loads(offline_installation_types)
                except json.JSONDecodeError:
                    offline_installation_types = []

            if not isinstance(offline_installation_types, list):
                offline_installation_types = []

            installation_type_display = ', '.join([str(x) for x in offline_installation_types if x])

            report_datetime_raw = report_data.get('report_datetime')
            report_timestamp = None
            if report_datetime_raw:
                try:
                    report_timestamp = datetime.fromisoformat(report_datetime_raw)
                except ValueError:
                    report_timestamp = None

            report = Report(
                user_id=current_user.id,
                report_type=offline_type,
                customer_name=report_data.get('customer_name'),
                company_project=report_data.get('company_project'),
                installation_type=installation_type_display if offline_type == 'installation' else None,
                installation_types=(
                    json.dumps(offline_installation_types, ensure_ascii=False)
                    if offline_type == 'installation'
                    else None
                ),
                protections_count=offline_protections_count if offline_type == 'installation' else None,
                address=report_data.get('address'),
                status=report_data.get('status'),
                notes=report_data.get('notes', ''),
                timestamp=report_timestamp or datetime.fromisoformat(report_data.get('timestamp')) if report_data.get('timestamp') else datetime.utcnow()
            )
            db.session.add(report)
            db.session.flush()

            # Add products
            for product in report_data.get('products', []):
                if product.get('name') and product.get('quantity'):
                    unit = product.get('unit') or 'unit'
                    if unit not in ['unit', 'meter']:
                        unit = 'unit'

                    report_product = ReportProduct(
                        report_id=report.id,
                        product_name=product['name'],
                        quantity=float(product['quantity']),
                        quantity_unit=unit
                    )
                    db.session.add(report_product)

            synced_count += 1

        except Exception as e:
            errors.append(str(e))

    db.session.commit()

    return jsonify({
        'success': True,
        'synced_count': synced_count,
        'errors': errors
    })

@app.route('/uploads/reports/<path:filename>')
@login_required
def uploaded_file(filename):
    """Serve uploaded files"""
    return send_from_directory(Config.UPLOAD_FOLDER, filename)

# PWA routes
@app.route('/manifest.json')
def manifest():
    return send_from_directory('static', 'manifest.json')

@app.route('/sw.js')
def service_worker():
    response = send_from_directory('static/js', 'sw.js')
    response.headers['Service-Worker-Allowed'] = '/'
    response.headers['Content-Type'] = 'application/javascript'
    return response

# Error handlers
@app.errorhandler(404)
def not_found(e):
    if request.is_json:
        return jsonify({'error': 'לא נמצא'}), 404
    return render_template('404.html'), 404

@app.errorhandler(500)
def server_error(e):
    if request.is_json:
        return jsonify({'error': 'שגיאת שרת'}), 500
    return render_template('500.html'), 500

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
